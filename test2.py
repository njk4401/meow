import pandas as pd
import time
import logging
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import re

# --- Setup logging ---
logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

start_time = time.time()

# --- Load IMDb datasets ---
logger.info("Loading title.basics.tsv.gz...")
usecols_basics = ["tconst", "titleType", "primaryTitle", "originalTitle",
                  "startYear", "runtimeMinutes", "genres"]
basics = pd.read_csv("title.basics.tsv.gz", sep="\t", na_values="\\N",
                     compression="gzip", usecols=usecols_basics, low_memory=False)
logger.info(f"Loaded {len(basics):,} titles from basics dataset.")


logger.info(f"Splitting sections...")
types = {
    "movie": "Movies",
    "tvMovie": "TVMovies",
    "tvSeries": "Series",
    "tvEpisode": "Episodes",
    "tvSpecial": "Specials",
    "tvMiniSeries": "MiniSeries",
    "short": "Shorts",
    "video": "Videos",
    "videoGame": "VideoGames"
}

dfs = {}
for ttype, name in types.items():
    df = basics[basics["titleType"] == ttype].copy()
    logger.info(f"{len(df):,} {name.lower()}")

    # Build IMDb hyperlink
    df["IMDb Link"] = "https://www.imdb.com/title/" + df["tconst"]

    # Rename and select columns
    df = df.rename(columns={
        "primaryTitle": "Title",
        "runtimeMinutes": "Length",
        "startYear": "Year",
        "genres": "Genres",
    })[["Title", "Length", "Year", "Genres", "IMDb Link"]]

    dfs[name] = df  # store in dictionary for later use

# --- Write all sheets to Excel ---
with pd.ExcelWriter("imdb.xlsx", engine="openpyxl") as writer:
    for sheet_name, df in dfs.items():
        if len(df) > 1048576:
            logger.info(f'Skipping {sheet_name} (too large)')
            continue
        logger.info(f'Writing {sheet_name}...')
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # --- Make titles clickable ---
        imdb_col_idx = df.columns.get_loc("IMDb Link") + 1  # 1-indexed in openpyxl
        for row in range(2, len(df) + 2):
            ws[f"A{row}"].hyperlink = ws.cell(row=row, column=imdb_col_idx).value
            ws[f"A{row}"].style = "Hyperlink"

        # --- Delete IMDb Link column ---
        ws.delete_cols(imdb_col_idx)

        # --- Create Excel Table ---
        last_row = len(df) + 1
        last_col = len(df.columns) - 1  # minus 1 because IMDb Link deleted
        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName=f"{sheet_name}Table", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)

        # --- Auto-adjust column widths ---
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 80)

logger.info(f"Saved spreadsheet in {time.time() - start_time:.2f} seconds")
