import os
import time
import logging
from collections.abc import Generator, Sequence

import requests
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


START_TIME = int(time.time())

type JSON = JSONArray | JSONObject
type JSONArray = list[JSONArray | JSONObject | JSONPrimitive]
type JSONObject = dict[str, JSONArray | JSONObject | JSONPrimitive]
type JSONPrimitive = bool | float | str

IMDB_DATA_URL = 'https://datasets.imdbws.com/'
FILES = ('title.basics.tsv.gz', 'title.ratings.tsv.gz')
BASICS_COLS = (
    'tconst', 'titleType', 'primaryTitle',
    'startYear', 'runtimeMinutes', 'genres'
)
RATINGS_COLS = ('tconst', 'averageRating', 'numVotes')

API = 'https://api.imdbapi.dev/titles:batchGet?'
LOCAL_CACHE = 'local_data.csv'
LOCAL_COLS = ('tconst', 'originCountry')

# --- Setup logging ---
logging.basicConfig(level=logging.INFO, format='[%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)


def is_fresh(file: os.PathLike, max_age: float) -> bool:
    """Return True if a file exists and is fresh.

    Parameters:
        file (PathLike):
            File path string.
        max_age (float):
            The maximum age of `file` to be considered "fresh" in seconds.
    """
    if not os.path.exists(file):
        return False
    age = time.time() - os.path.getmtime(file)
    return age < max_age


def download_imdb_data(files: Sequence[str] = FILES) -> None:
    """Download IMDb datasets if missing or outdated.

    Parameters:
        files (Sequence[str]):
            Sequence of IMDB dataset file names to download.
    """
    for file in files:
        # Skip if the file is less than a day old
        if is_fresh(file, max_age=24*3600):
            logger.info(f'Skipping "{file}" download')
            continue

        logger.info(f'Downloading {file}...')
        url = IMDB_DATA_URL+file
        response = requests.get(url, stream=True)
        response.raise_for_status()

        with open(file, 'wb') as f:
            for chunk in response.iter_content(8192):
                f.write(chunk)

        logger.info(f'Downloaded "{file}"')


def chunks(lst: Sequence, n: int) -> Generator[Sequence]:
    """Separate a sequence into smaller sequences of `n` items."""
    for i in range(0, len(lst), n):
        yield lst[i:i+n]


def fetch(url: str, timeout: float = 15,
          retries: int = 5, delay: float = 2) -> JSON | None:
    """Fetch JSON data with retry in case of rate limiting.

    Parameters:
        url (str):
            URL to fetch.
        timeout (float):
            Fetch timeout in seconds.
        retries (int):
            Number of retries to attempt after being rate limited.
        delay (float):
            Initial time between retries in seconds.
            Will be scaled by a factor of 2 on each retry.

    Returns:
        response (JSON | None):
            JSON response if fetch was successful, otherwise None.
    """
    for attempt in range(retries):
        try:
            resp = requests.get(url, timeout=timeout)
            # If being rate limitted, wait and retry
            if resp.status_code == 429:
                wait = delay * (2**attempt)
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        # If another error raised, wait and retry
        except requests.RequestException:
            wait = delay * (2**attempt)
            logger.warning(f'Request failed - Retrying in {wait}s...')
            time.sleep(wait)

    logger.error(f'Failed after {retries} retries: {url}')
    return None


def load_cache(file: os.PathLike = LOCAL_CACHE) -> pd.DataFrame:
    """Load additional data from local CSV if exists.

    Parameters:
        file (PathLike):
            Local CSV file to attempt to load from.

    Returns:
        df:
            Pandas DataFrame of the CSV content.
    """
    if not os.path.exists(file):
        return pd.DataFrame(columns=LOCAL_COLS)
    df = pd.read_csv(file, usecols=LOCAL_COLS)
    logger.info(f'Loaded {len(df):,} local entries')
    return df


def save_cache(df: pd.DataFrame, file: os.PathLike = LOCAL_CACHE) -> None:
    """Save additional data to a local CSV.

    Parameters:
        df (DataFrame):
            Pandas DataFrame of the CSV content.
        file (PathLike):
            Local CSV file to save to.
    """
    df.to_csv(file, index=False)
    logger.info(f'Saved {len(df):,} local country origins')


def apply_base_filters(
        df: pd.DataFrame, *,
        titles: Sequence[str] = (),
        ratings: Sequence[float] = (),
        min_votes: int = 0,
    ) -> pd.DataFrame:
    """Filter IMDb title dataset.

    Parameters:
        df (DataFrame):
            IMDb basics title dataset as a Pandas DataFrame.
        titles (Sequence[str]):
            Types of titles to include.
        ratings (Sequence[float]):
            Sequence of min, max rating value.
        min_votes (int):
            Minimum number of rating votes.
    """
    VALID_TITLES = {
        'movie', 'tvMovie', 'tvSeries', 'tvEpisode', 'tvSpecial',
        'tvMiniSeries', 'short', 'video', 'videoGame'
    }

    if titles:
        if invalid := set(titles).difference(VALID_TITLES):
            logger.warning(
                f'Invalid title_type value(s): "{','.join(invalid)}", '
                f'Valid: {VALID_TITLES}'
            )
        df = df[df['titleType'].isin(set(titles))]
        logger.info(
            f'{len(df):,} titles after filtering for {','.join(titles)}'
        )
    if ratings:
        try:
            min_rating, max_rating = ratings[:2]
            if (float(min_rating) < 0) or (float(max_rating) > 0):
                logger.warning('Rating range exceeds tolerable range: [1, 10]')
            df = df[df[ratings[0] <= df['averageRating'] <= ratings[1]]]
            logger.info(
                f'{len(df):,} titles with ratings between {'-'.join(ratings)}'
            )
        except (IndexError, TypeError, ValueError):
            logger.error(
                'Ratings must be a sequence of numeric (min, max) elements'
            )
        except KeyError:
            logger.error('The dataset does not include ratings')
    if min_votes:
        try:
            if int(min_votes) < 0:
                logger.warning('Vote count of less than 0 will have no effect')
            df = df[df['numVotes'] >= min_votes]
            logger.info(f'{len(df):,} titles with at least {min_votes} votes')
        except TypeError:
            logger.error('Vote count must be numeric')
        except KeyError:
            logger.error('The dataset does not include votes')

    return df


def insert_cache_data(df: pd.DataFrame) -> pd.DataFrame:
    """Insert additional data into IMDb dataset."""
    logger.info('Generating additional criteria...')
    cache = load_cache()
    existing = set(cache['tconst'])
    tconsts = df['tconst'].tolist()

    pending = [t for t in tconsts if t not in existing]
    new_entries = []
    for i, batch in enumerate(chunks(pending, 5), start=1):
        print(f'{i*5}/{len(pending)}', end='\r', flush=True)
        ids = '&'.join(f'titleIds={b}' for b in batch)
        data: JSONObject = fetch(API+ids)
        if not data:
            continue

        titles: JSONArray = data.get('titles', [])
        for item in titles:
            if not (tid := item.get('id')):
                continue
            if not (countries := item.get('originCountries', [])):
                country = 'N/A'
            else:
                country = countries[0].get('name', 'N/A').split('(')[0]

            new_entries.append(dict(tconst=tid, originCountry=country))

        # Save every 100 requests
        if i % 100 == 0 and new_entries:
            new = pd.DataFrame(new_entries)
            cache = pd.concat(
                [cache, new], ignore_index=True
            ).drop_duplicates('tconst')
            save_cache(cache)
            new_entries = []

        time.sleep(0.1)

    if new_entries:
        new = pd.DataFrame(new_entries)
        cache = pd.concat(
            [cache, new], ignore_index=True
        ).drop_duplicates('tconst')
        save_cache(cache)

    df = df.merge(cache, on='tconst', how='left')

    return df


def apply_additional_filters(
        df: pd.DataFrame, *,
        countries: Sequence[str] = ()
    ) -> pd.DataFrame:
    """Filter IMDb title dataset with additional data.

    Parameters:
        df (DataFrame):
            IMDb basics title dataset as a Pandas DataFrame.
        countries (Sequence[str]):
            Title origin countries to include.
    """
    if countries:
        try:
            df = df[df['originCountry'] in set(countries)]
            logger.info(
                f'{len(df):,} titles after filtering for {','.join(countries)}'
            )
        except KeyError:
            logger.error('Dataset does not include origin countries')

    return df


def to_excel(df: pd.DataFrame) -> None:
    """Write DataFrame to Excel workbook."""
    logger.info('Writing Excel...')
    with pd.ExcelWriter('imdb.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All')
        ws = writer.sheets['All']

        # --- Make titles clickable ---
        imdb_col = df.columns.get_loc('IMDb Link')+1
        for row in range(2, len(df)+2):
            ws[f'A{row}'].hyperlink = ws.cell(row=row, column=imdb_col).value
            ws[f'A{row}'].style = 'Hyperlink'

        # --- Delete IMDb Link column ---
        ws.delete_cols(imdb_col)

        # --- Create Excel Table ---
        last_row = len(df)+1
        last_col = len(df.columns)-1
        table_ref = f'A1:{get_column_letter(last_col)}{last_row}'
        table = Table(displayName='IMDbTable', ref=table_ref)
        style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)

        # --- Auto-adjust column widths ---
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_len+2, 80
            )


def main(base_filters={}, additional_filters={}) -> None:
    """Create spreadsheet of IMDb titles."""
    download_imdb_data()

    # --- Load IMDb datasets ---
    logger.info('Loading titles...')
    basics_list = []
    chunks = pd.read_csv('title.basics.tsv.gz', sep='\t', na_values='\\N',
                         compression='gzip', usecols=BASICS_COLS,
                         chunksize=50_000)
    for chunk in chunks:
        chunk = chunk[chunk['titleType'] == 'movie']
        basics_list.append(chunk)
    basics = pd.concat(basics_list, ignore_index=True)
    logger.info(f'Loaded {len(basics):,} titles from basics dataset')
    ratings = pd.read_csv(
        'title.ratings.tsv.gz', sep='\t', na_values='\\N', compression='gzip',
        usecols=RATINGS_COLS
    )
    df = basics.merge(ratings, on='tconst', how='inner')
    logger.info(f'{len(df):,} titles with ratings')

    df = apply_base_filters(df, **base_filters)

    df = insert_cache_data(df)
    df = apply_additional_filters(df, **additional_filters)

    # --- Build IMDb hyperlink column ---
    df['IMDb Link'] = 'https://www.imdb.com/title/'+df['tconst']

    # --- Rename and select columns ---
    df = df.rename(columns={
        'genres': 'Genres',
        'numVotes': 'Votes',
        'startYear': 'Year',
        'primaryTitle': 'Title',
        'averageRating': 'Rating',
        'originCountry': 'Country',
        'runtimeMinutes': 'Runtime'
    })[[
        'Title', 'Rating', 'Year', 'Runtime', 'Votes',
        'Country', 'Genres', 'IMDb Link'
    ]]

    # -- Ensure Runtime is numeric --
    df['Runtime'] = pd.to_numeric(df['Runtime'], errors='coerce')

    # --- Write to Excel ---
    to_excel(df)

    END_TIME = int(time.time()) - START_TIME
    secs = END_TIME % 60
    mins = (END_TIME // 60) % 60
    hours = END_TIME // 3600
    logger.info(f'Saved {len(df):,} titles in {hours}:{mins:02}:{secs:02}')


if __name__ == '__main__':
    main(
        base_filters=dict(
            titles={'movie'}, min_votes=1000
        )
    )
